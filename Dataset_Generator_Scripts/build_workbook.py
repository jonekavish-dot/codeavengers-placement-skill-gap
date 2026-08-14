import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

NAVY = "12233F"
WHITE = "FFFFFF"
LIGHT = "F4F5F7"

students = pd.read_csv("Students_Placement.csv")
skills = pd.read_csv("Student_Skills.csv")
demand = pd.read_csv("Industry_Skill_Demand.csv")

wb = Workbook()

def style_sheet(ws, df, table_name, currency_cols=None, pct_cols=None):
    currency_cols = currency_cols or []
    pct_cols = pct_cols or []
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))

    header_font = Font(name="Arial", bold=True, color=WHITE, size=10.5)
    header_fill = PatternFill("solid", fgColor=NAVY)
    body_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D6DAE2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_rows = ws.max_row
    n_cols = ws.max_column
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r in range(2, n_rows + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.border = border
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
            col_name = df.columns[c - 1]
            if col_name in currency_cols:
                cell.number_format = '#,##0.00'
            if col_name in pct_cols:
                cell.number_format = '0"%"'

    # column widths
    for c, col_name in enumerate(df.columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).values])
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 3, 10), 32)

    ws.freeze_panes = "A2"
    ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
    ws.add_table(tbl)

# ---- Read Me sheet (first) ----
ws0 = wb.active
ws0.title = "Read Me"
ws0.sheet_view.showGridLines = False
title_font = Font(name="Arial", bold=True, size=14, color=NAVY)
label_font = Font(name="Arial", bold=True, size=10.5)
body_font = Font(name="Arial", size=10.5)
warn_fill = PatternFill("solid", fgColor="FBE9E7")

ws0["A1"] = "PLACEMENT SKILL-GAP HUB — SAMPLE / DEMONSTRATION DATASET"
ws0["A1"].font = title_font
ws0.merge_cells("A1:F1")

ws0["A3"] = ("This is synthetic demonstration data generated for the HACKORBIT 2K26 build. "
             "No real Bannari Amman Institute of Technology student, placement, or recruiter "
             "data is used anywhere in this workbook.")
ws0["A3"].font = Font(name="Arial", italic=True, size=10.5, color="C0392B")
ws0.merge_cells("A3:F4")
ws0["A3"].alignment = Alignment(wrap_text=True, vertical="top")
for row in ws0["A3:F4"]:
    for cell in row:
        cell.fill = warn_fill

rows_info = [
    ("", "", 18),
    ("Table", "Contents", 18),
    ("Students_Placement", "One row per student — Branch, Batch, Placement_Status, Company, Package", 28),
    ("Student_Skills", "One row per student-skill pair — Assessment_Score (0–100) as a readiness proxy", 28),
    ("Industry_Skill_Demand", "One row per skill — Demand_Level and Companies_Requiring_Count", 28),
    ("", "", 12),
    ("How the numbers were built (so they hold up to questions):", "", 20),
    ("Readiness", "Varies by branch — the skill set here (Python/SQL/Power BI/ML/Java) is "
                   "software-and-analytics focused, which structurally favors CSE/AIDS/IT over "
                   "ECE/MECH. This is a labeled limitation of the demo skill list, not a claim "
                   "about branch ability.", 62),
    ("Industry Demand", "Derived directly from how many of the 22 sample companies require each "
                         "skill — Demand_Level and Companies_Requiring_Count come from the same "
                         "underlying assignment, not two independently invented numbers.", 55),
    ("Placement & Package", "Probability of placement and package level both correlate with a "
                             "student's composite technical-skill readiness, with randomness layered "
                             "on top — not a fixed lookup.", 48),
]
r = 6
for label, val, height in rows_info:
    ws0.cell(row=r, column=1, value=label).font = label_font if val or label.endswith(":") else body_font
    ws0.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    c2 = ws0.cell(row=r, column=2, value=val)
    c2.font = body_font
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws0.row_dimensions[r].height = height
    r += 1

ws0.column_dimensions["A"].width = 24
for col in "BCDEFG":
    ws0.column_dimensions[col].width = 15

# ---- Data sheets ----
ws1 = wb.create_sheet("Students_Placement")
style_sheet(ws1, students, "TblStudentsPlacement", currency_cols=["Package"])

ws2 = wb.create_sheet("Student_Skills")
style_sheet(ws2, skills, "TblStudentSkills")

ws3 = wb.create_sheet("Industry_Skill_Demand")
style_sheet(ws3, demand, "TblIndustryDemand")

wb.save("Placement_SkillGap_Hub_Dataset.xlsx")
print("Workbook saved.")
